import pandas as pd
import openpyxl
from flask import Flask, render_template, request, jsonify

app = Flask(__name__)

class CourseEditor:
    def __init__(self, file_path="course_data.xlsx"):
        self.file_path = file_path
        self.COURSES_DATA = [
            'day', 'time', 'course_id', 'course_name', 'teacher',
            'location', 'extra_info', 'grades', 'credits',
            'max_capacity', 'enrolled'
        ]
        self.STUDENTS_COLUMNS = ['student_id', 'course_id']  # 假設學生資料的欄位

    def load_data(self, sheet_name):
        """載入Excel數據"""
        try:
            df = pd.read_excel(self.file_path, sheet_name=sheet_name)
            if sheet_name == 'courses':
                df.columns = self.COURSES_DATA
            elif sheet_name == 'students':
                df.columns = self.STUDENTS_COLUMNS
            return df
        except Exception as e:
            print(f"Error loading data from {sheet_name}: {e}")
            return pd.DataFrame()

    def save_data(self, courses_df, students_df):
    """保存課程和學生數據，同時保留其他表單"""
    try:
        # 載入整個 Excel 文件
        workbook = openpyxl.load_workbook(self.file_path)

        # 保存原有的表單（保留 'students' 表單）
        if 'courses' in workbook.sheetnames:
            del workbook['courses']
        
        # 使用 ExcelWriter 存回文件，保證兩個表單都被更新
        with pd.ExcelWriter(self.file_path, engine='openpyxl') as writer:
            writer.book = workbook  # 使用原來的工作簿
            courses_df.to_excel(writer, sheet_name='courses', index=False)  # 保存課程表
            students_df.to_excel(writer, sheet_name='students', index=False)  # 保存學生表

    except Exception as e:
        print(f"Error saving data: {e}")



    def delete_course(self, course_id):
    """刪除課程，並從學生資料中刪除選修該課程的學生"""
    try:
        # 載入課程和學生數據
        courses_df = self.load_data('courses')
        students_df = self.load_data('students')
        
        # 檢查課程是否存在
        if course_id not in courses_df['course_id'].values:
            return False, "課程不存在"
        
        # 獲取課程信息
        course_info = courses_df[courses_df['course_id'] == course_id].iloc[0]

        # 刪除 'courses' 表單中的課程
        courses_df = courses_df[courses_df['course_id'] != course_id]

        # 只刪除 'students' 表單中選修該課程的學生資料
        students_df = students_df[students_df['course_id'] != course_id]

        # 使用 openpyxl 打開 Excel 檔案
        workbook = openpyxl.load_workbook(self.file_path)
        
        # 保存更新後的資料，注意不會刪除其他表單
        with pd.ExcelWriter(self.file_path, engine='openpyxl') as writer:
            writer.book = workbook  # 使用原來的工作簿
            # 更新 courses 和 students 表單
            courses_df.to_excel(writer, sheet_name='courses', index=False)
            students_df.to_excel(writer, sheet_name='students', index=False)
        
        return True, f"課程 {course_id} ({course_info['course_name']}) 已成功刪除"
    
    except Exception as e:
        print(f"Error deleting course: {e}")
        return False, f"刪除失敗: {str(e)}"




@app.route('/delete_course', methods=['POST'])
def delete_course_route():
    """刪除課程的API路由"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "message": "無效的請求數據"}), 400

        course_id = data.get('course_id')
        if not course_id:
            return jsonify({"success": False, "message": "缺少課程ID"}), 400
        
        success, message = course_editor.delete_course(course_id)
        return jsonify({"success": success, "message": message})

    except Exception as e:
        return jsonify({"success": False, "message": f"處理請求時發生錯誤: {str(e)}"}), 500



@app.route('/course_management')
def course_management():
    """課程管理頁面路由"""
    courses = course_editor.get_all_courses()
    return render_template('course-management.html', courses=courses)


@app.route('/get_course/<course_id>')
def get_course_route(course_id):
    """獲取特定課程信息的API路由"""
    df = course_editor.load_data('courses')
    course = df[df['course_id'] == course_id].to_dict('records')
    if course:
        return jsonify(course[0])
    return jsonify({"error": "課程不存在"}), 404


@app.route('/edit_course', methods=['POST'])
def edit_course_route():
    """編輯課程的API路由"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "message": "無效的請求數據"}), 400

        course_id = data.get('course_id')
        updated_info = data.get('updated_info')
        
        if not course_id or not updated_info:
            return jsonify({"success": False, "message": "缺少必要參數"}), 400
        
        success, message = course_editor.edit_course(course_id, updated_info)
        return jsonify({"success": success, "message": message})

    except Exception as e:
        return jsonify({"success": False, "message": f"處理請求時發生錯誤: {str(e)}"}), 500


if __name__ == '__main__':
    course_editor = CourseEditor()
    app.run(debug=True)
